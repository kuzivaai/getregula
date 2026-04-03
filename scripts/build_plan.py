"""
Build Regula go-to-market plan spreadsheet.
Run: python3 scripts/build_plan.py
Output: docs/regula_gtm_plan.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
# from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1  # unused
from openpyxl.chart import BarChart, Reference
from datetime import date, timedelta
import calendar

wb = Workbook()

# ── Colour palette ──────────────────────────────────────────────
BLACK   = "1A1A2E"
NAVY    = "16213E"
BLUE    = "0F3460"
ACCENT  = "E94560"
WHITE   = "FFFFFF"
LGRAY   = "F4F6F8"
MGRAY   = "DEE2E6"
DGRAY   = "6C757D"
GREEN   = "28A745"
AMBER   = "FFC107"
RED_    = "DC3545"
TEAL    = "17A2B8"
PURPLE  = "6F42C1"
ORANGE  = "FD7E14"

def fill(hex_): return PatternFill("solid", fgColor=hex_)
def font(hex_, bold=False, sz=11, italic=False):
    return Font(color=hex_, bold=bold, size=sz, italic=italic, name="Calibri")
def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def border_thin(sides="bottom"):
    s = Side(style="thin", color=MGRAY)
    b = Border()
    if "top"    in sides: b.top    = s
    if "bottom" in sides: b.bottom = s
    if "left"   in sides: b.left   = s
    if "right"  in sides: b.right  = s
    if "all"    in sides:
        b.top = b.bottom = b.left = b.right = s
    return b

def header_row(ws, row, cells, bg=NAVY, fg=WHITE, sz=11, bold=True, height=22):
    ws.row_dimensions[row].height = height
    for col, val in enumerate(cells, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill(bg)
        c.font = font(fg, bold=bold, sz=sz)
        c.alignment = align("center", "center")
        c.border = border_thin("all")

def data_row(ws, row, cells, bg=WHITE, fg=BLACK, bold=False, wrap=False, height=None):
    if height: ws.row_dimensions[row].height = height
    for col, val in enumerate(cells, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill(bg)
        c.font = font(fg, bold=bold)
        c.alignment = align("left", "center", wrap=wrap)
        c.border = border_thin("bottom")

def merge_title(ws, row, start_col, end_col, text, bg=BLACK, fg=WHITE, sz=14, bold=True, height=32):
    ws.row_dimensions[row].height = height
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=start_col, value=text)
    c.fill = fill(bg)
    c.font = font(fg, bold=bold, sz=sz)
    c.alignment = align("center", "center")

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def badge(ws, row, col, text, bg, fg=WHITE):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill(bg)
    c.font = font(fg, bold=True, sz=9)
    c.alignment = align("center", "center")
    c.border = border_thin("all")


# ═══════════════════════════════════════════════════════════════
# TAB 1 — VISION, MISSION & STRATEGIC OKRs
# ═══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "1 · Vision & OKRs"
ws1.sheet_view.showGridLines = False

merge_title(ws1, 1, 1, 7, "REGULA  ·  GO-TO-MARKET STRATEGIC PLAN", BLACK, WHITE, 16)
merge_title(ws1, 2, 1, 7, "April 2026 – March 2027  ·  EU AI Act High-Risk Deadline: 2 August 2026", BLUE, WHITE, 13)

# Vision / Mission
ws1.row_dimensions[4].height = 14
ws1.merge_cells("A4:G4")
ws1.cell(4, 1).fill = fill(LGRAY)

for row, label, text, bg in [
    (5, "VISION",
     "Become the default developer tool for EU AI Act compliance — the first thing a team runs before shipping AI to Europe.",
     NAVY),
    (7, "MISSION",
     "Give every developer — from solo founder to enterprise team — a free, honest, code-level answer to 'is my AI system high-risk?' in under 60 seconds.",
     BLUE),
    (9, "PRODUCT",
     "Regula v1.5.0 · Free CLI · PyPI: regula-ai · GitHub: kuzivaai/getregula · getregula.com · MIT licence · Zero dependencies",
     TEAL),
]:
    ws1.row_dimensions[row].height = 38
    ws1.merge_cells(f"A{row}:B{row}")
    ws1.merge_cells(f"C{row}:G{row}")
    c_label = ws1.cell(row, 1, label)
    c_label.fill = fill(bg)
    c_label.font = font(WHITE, bold=True, sz=11)
    c_label.alignment = align("center", "center")
    c_text = ws1.cell(row, 3, text)
    c_text.fill = fill(LGRAY)
    c_text.font = font(BLACK, sz=10)
    c_text.alignment = align("left", "center", wrap=True)
    ws1.row_dimensions[row+1].height = 6

# Separator
ws1.row_dimensions[11].height = 10

# Market context
merge_title(ws1, 12, 1, 7, "MARKET CONTEXT (verified, April 2026)", DGRAY, WHITE, 12)
context_rows = [
    ("AI governance market size 2026",      "$492M (Gartner Feb 2026)",          "Growing to $1B+ by 2030"),
    ("EU AI Act high-risk deadline",        "2 August 2026 (Annex III)",         "Digital Omnibus may extend to Dec 2027 — treat Aug as binding"),
    ("Enterprise GRC platforms (OneTrust, IBM)", "$50,000+/year (est.)",          "Regula is free. Gap is real. KLA Digital SaaS starts €299/mo."),
    ("SME self-assessment cost",            "€9,500 – €14,500 + internal time",  "Regula eliminates the triage phase"),
    ("Max penalty",                         "€35M or 7% global turnover",        "Exceeds GDPR — real compliance urgency"),
    ("Top open-source competitor",          "EuConform — 107 stars (Dec 2025)",  "Browser-based, no CLI, no CI/CD integration"),
    ("Regula GitHub stars at launch",       "0 (repo created 25 Mar 2026)",      "Honest starting point"),
    ("Regula PyPI",                         "regula-ai v1.5.0 — live",           "748 tests passing, 0 open issues"),
]
header_row(ws1, 13, ["Metric", "Data Point", "Implication"], NAVY, WHITE)
for i, (m, d, imp) in enumerate(context_rows):
    bg = LGRAY if i % 2 == 0 else WHITE
    data_row(ws1, 14+i, [m, d, imp], bg=bg, wrap=True, height=22)

ws1.row_dimensions[23].height = 10

# OKRs header
merge_title(ws1, 24, 1, 7, "ANNUAL OKRs  ·  April 2026 – March 2027", ACCENT, WHITE, 13)
header_row(ws1, 25, ["#", "Objective", "Key Result", "Target", "By", "Measure", "Status"], NAVY, WHITE)

okrs = [
    # Awareness OKRs
    ("O1", "Build developer awareness",    "GitHub stars",                 "500",       "Oct 2026", "gh api repos/kuzivaai/getregula --jq .stargazers_count",  "Not started"),
    ("O1", "Build developer awareness",    "PyPI monthly downloads",       "1,000/mo",  "Oct 2026", "pypistats.org/packages/regula-ai",                        "Not started"),
    ("O1", "Build developer awareness",    "HN Show HN front page",        "1 post",    "May 2026", "news.ycombinator.com",                                    "Draft ready"),
    ("O1", "Build developer awareness",    "DEV.to article published",     "1 article", "May 2026", "dev.to/@kuzivaai",                                        "Not started"),
    # Adoption OKRs
    ("O2", "Drive real-world adoption",    "Unique CLI installs (pip)",     "500",       "Dec 2026", "PyPI download stats",                                     "Not started"),
    ("O2", "Drive real-world adoption",    "GitHub issues opened by users", "10",        "Oct 2026", "github issues list",                                      "Not started"),
    ("O2", "Drive real-world adoption",    "Public 'regula check' reports", "5",         "Sep 2026", "Link collection",                                         "Not started"),
    # Ecosystem OKRs
    ("O3", "Enter compliance ecosystem",   "Listed on artificialintelligenceact.eu", "1", "Jul 2026", "Manual verification",                                  "Not started"),
    ("O3", "Enter compliance ecosystem",   "Listed in awesome-claude-code PR merged", "1","Jun 2026", "github PR",                                             "Not started"),
    ("O3", "Enter compliance ecosystem",   "IAPP or EU SME Alliance mention",  "1",     "Aug 2026", "Manual tracking",                                         "Not started"),
    # Product OKRs
    ("O4", "Maintain product quality",     "Test pass rate",                "100%",      "Ongoing",  "python3 tests/test_classification.py",                    "748/748 PASS"),
    ("O4", "Maintain product quality",     "PyPI release cadence",          "1/quarter", "Ongoing",  "pypi.org/project/regula-ai",                              "v1.5.0 live"),
    ("O4", "Maintain product quality",     "False positive rate on real OSS","< 5%",     "Sep 2026", "Manual evaluation on 3+ repos",                           "Not measured"),
    # Revenue (if applicable)
    ("O5", "Explore sustainability",       "Enterprise enquiries received", "3",         "Mar 2027", "Email / contact form",                                    "Not started"),
    ("O5", "Explore sustainability",       "Open-source sponsors (GitHub)", "1",         "Mar 2027", "github.com/sponsors/kuzivaai",                            "Not started"),
]

okr_colors = {
    "O1": ("E8F4FD", BLUE),
    "O2": ("E8F8F0", GREEN),
    "O3": ("FFF8E8", ORANGE),
    "O4": ("F3E8FF", PURPLE),
    "O5": ("FFE8E8", RED_),
}

status_colors = {
    "Not started": MGRAY,
    "Draft ready": AMBER,
    "748/748 PASS": GREEN,
    "v1.5.0 live": GREEN,
    "Not measured": AMBER,
}

for i, (obj_id, obj, kr, target, by, measure, status) in enumerate(okrs):
    row = 26 + i
    ws1.row_dimensions[row].height = 20
    bg_row, _ = okr_colors[obj_id]
    data_row(ws1, row, [obj_id, obj, kr, target, by, measure, ""], bg=bg_row)
    badge(ws1, row, 7, status,
          status_colors.get(status, MGRAY),
          WHITE if status not in (MGRAY,) else BLACK)

set_col_widths(ws1, [5, 26, 38, 12, 11, 38, 14])

for r in range(1, 60):
    for c in range(1, 8):
        cell = ws1.cell(r, c)
        if cell.value is None:
            cell.fill = fill(WHITE)


# ═══════════════════════════════════════════════════════════════
# TAB 2 — 12-MONTH MILESTONE ROADMAP
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("2 · 12-Month Roadmap")
ws2.sheet_view.showGridLines = False

merge_title(ws2, 1, 1, 9, "12-MONTH MILESTONE ROADMAP  ·  April 2026 – March 2027", BLACK, WHITE, 16)
merge_title(ws2, 2, 1, 9, "Key external date: EU AI Act Annex III enforcement — 2 August 2026", ACCENT, WHITE, 13)

header_row(ws2, 3, ["Phase", "Month", "Milestone", "Actions Required", "Owner", "Success Metric", "Status", "Priority", "Notes"], NAVY, WHITE, height=24)

phases = [
    # ─ Phase 1: Pre-launch (Apr)
    ("Phase 1\nPre-Launch", "Apr 2026",
     "GitHub repo ready for HN",
     "✓ Description set\n✓ Topics set\n✓ Release v1.5.0 published\n✓ Homepage URL set\n□ README hero GIF/screenshot\n□ CONTRIBUTING.md updated",
     "You", "Repo looks polished", "In progress", "P0",
     "All ✓ items done today"),
    ("Phase 1\nPre-Launch", "Apr 2026",
     "Landing page verified live",
     "□ Confirm getregula.com serves index.html\n□ Check robots.txt, sitemap indexed\n□ Test email capture form end-to-end\n□ Mobile render check",
     "You", "getregula.com loads < 2s", "Not started", "P0",
     "Blocker before Show HN"),
    ("Phase 1\nPre-Launch", "Apr 2026",
     "DEV.to article drafted",
     "□ Title: 'I scanned 5 open-source AI repos for EU AI Act compliance'\n□ Run regula on LangChain, Hugging Face, PrivateGPT, Ollama, AutoGPT\n□ Show real output, honest false-positive rate\n□ Link to GitHub repo",
     "You", "Article published on DEV.to", "Not started", "P1",
     "Evidence-first marketing"),

    # ─ Phase 2: Launch (May)
    ("Phase 2\nLaunch", "May 2026",
     "Show HN submitted",
     "□ Post on Monday or Tuesday morning ET\n□ Title: Option A or C from SHOW_HN_DRAFT.md\n□ Link to GitHub repo, not landing page\n□ Be present for 4+ hours to answer comments\n□ Answer every technical question honestly",
     "You", "Top 30 on HN front page", "Not started", "P0",
     "Primary launch channel"),
    ("Phase 2\nLaunch", "May 2026",
     "awesome-claude-code PR submitted",
     "□ Find the awesome-claude-code repo\n□ Regula has Claude Code hooks in hooks/\n□ Submit PR with brief description",
     "You", "PR merged", "Not started", "P1",
     "Niche but highly targeted"),
    ("Phase 2\nLaunch", "May 2026",
     "Product Hunt submission",
     "□ Create PH product page with screenshots\n□ Schedule for Tuesday launch\n□ Ask 5 users to upvote on launch day",
     "You", "Top 10 in dev tools category", "Not started", "P2",
     "Second wave after HN"),

    # ─ Phase 3: Compliance ecosystem (Jun-Jul)
    ("Phase 3\nEcosystem", "Jun 2026",
     "Listed on artificialintelligenceact.eu",
     "□ Contact the site editors\n□ Submit tool to their tools/resources directory\n□ Offer to write a guest article on code-level risk detection",
     "You", "Listed on site (150K+ users/mo)", "Not started", "P0",
     "Highest-ROI channel for this audience"),
    ("Phase 3\nEcosystem", "Jun 2026",
     "EU Digital SME Alliance outreach",
     "□ Find their tool submission form\n□ Submit Regula as free SME compliance tool\n□ Emphasise: free, no account, 60 seconds",
     "You", "Listed in SME directory", "Not started", "P1",
     "Target audience: SMEs building AI"),
    ("Phase 3\nEcosystem", "Jul 2026",
     "IAPP outreach",
     "□ Find relevant IAPP community / knowledge base\n□ Contribute article or tool listing\n□ Frame as: 'triage tool for DPOs reviewing AI systems'",
     "You", "Mention in IAPP resource", "Not started", "P2",
     "DPO/compliance officer audience"),
    ("Phase 3\nEcosystem", "Jul 2026",
     "LinkedIn content series begins",
     "□ 4 posts: 'What the EU AI Act requires from your code'\n□ 1 post per week, each with a specific Article\n□ End each with: 'regula check . will flag this in seconds'\n□ Target: CTOs, DPOs, AI engineers",
     "You", "100+ impressions/post avg", "Not started", "P2",
     "Slow burn, builds authority"),

    # ─ Phase 4: Deadline window (Aug)
    ("Phase 4\nDeadline", "Aug 2026",
     "EU AI Act enforcement day (Aug 2)",
     "□ Publish blog post: 'The deadline is here. Here's what to check first.'\n□ Pin to GitHub README\n□ Share on LinkedIn, HN, DEV.to\n□ Be available for DMs/emails from developers",
     "You", "500+ GitHub stars by Aug 15", "Not started", "P0",
     "Highest organic traffic window of the year"),
    ("Phase 4\nDeadline", "Aug 2026",
     "Real-world accuracy evaluation published",
     "□ Run regula on 5 public OSS AI repos\n□ Manually verify findings (no inflation)\n□ Publish honest precision/recall numbers\n□ Compare to self-test numbers",
     "You", "Published evaluation doc", "Not started", "P0",
     "Credibility is the product"),

    # ─ Phase 5: Post-launch growth (Sep-Dec)
    ("Phase 5\nGrowth", "Sep 2026",
     "Community feedback integration",
     "□ Review all GitHub issues opened by users\n□ Prioritise false-positive improvements\n□ Ship v1.6.0 with user-reported fixes\n□ Thank contributors publicly",
     "You", "v1.6.0 released", "Not started", "P1",
     "Issues are a success metric"),
    ("Phase 5\nGrowth", "Oct 2026",
     "1,000 PyPI downloads/month milestone",
     "□ Track via pypistats.org weekly\n□ If behind: revisit Show HN with new angle\n□ If ahead: write case study post",
     "You", "1,000 downloads/month", "Not started", "P1",
     "Lagging indicator of distribution success"),
    ("Phase 5\nGrowth", "Nov 2026",
     "r/python and r/programming posts",
     "□ Participate in existing compliance/AI Act threads for 4 weeks first\n□ Post 'I scanned the top 50 PyPI AI packages for EU AI Act compliance'\n□ Show real data, link to results gist",
     "You", "200+ upvotes on main post", "Not started", "P2",
     "80/20 rule: participate first"),
    ("Phase 5\nGrowth", "Dec 2026",
     "500 GitHub stars milestone",
     "□ Review star acquisition trajectory\n□ If behind: identify which channel drove most stars\n□ Double down on that channel",
     "You", "500 stars", "Not started", "P1",
     "Vanity metric but signals discoverability"),

    # ─ Phase 6: Sustainability (Jan-Mar 2027)
    ("Phase 6\nSustain", "Jan 2027",
     "GitHub Sponsors page live",
     "□ Enable GitHub Sponsors for kuzivaai org\n□ Set tiers: $5 (supporter), $25 (team), $100 (company)\n□ Add sponsor badge to README",
     "You", "1 sponsor", "Not started", "P2",
     "Sustainability, not revenue"),
    ("Phase 6\nSustain", "Feb 2027",
     "Enterprise enquiry pipeline",
     "□ Add 'Enterprise / team support' form to getregula.com\n□ Define offering: onboarding, custom rules, audit support\n□ Price range: €500-€2,000/day consulting",
     "You", "3 enquiries received", "Not started", "P2",
     "Explore only if open-source traction exists"),
    ("Phase 6\nSustain", "Mar 2027",
     "Year-1 retrospective",
     "□ Measure all OKRs vs targets\n□ Write public retrospective post\n□ Define v2 roadmap based on user feedback",
     "You", "Retro published", "Not started", "P3",
     "Accountability and direction"),
]

priority_colors = {"P0": RED_, "P1": ORANGE, "P2": AMBER, "P3": TEAL}
status_colors2 = {
    "In progress": AMBER,
    "Not started": MGRAY,
    "Done": GREEN,
}
phase_colors = {
    "Phase 1\nPre-Launch": "EBF5FB",
    "Phase 2\nLaunch":     "E8F8F0",
    "Phase 3\nEcosystem":  "FFF8E8",
    "Phase 4\nDeadline":   "FFE8E8",
    "Phase 5\nGrowth":     "F3E8FF",
    "Phase 6\nSustain":    "E8EAF6",
}

for i, (phase, month, milestone, actions, owner, metric, status, priority, notes) in enumerate(phases):
    row = 4 + i
    ws2.row_dimensions[row].height = 70
    bg = phase_colors.get(phase, WHITE)
    for col, val in enumerate([phase, month, milestone, actions, owner, metric, "", "", notes], 1):
        c = ws2.cell(row=row, column=col, value=val)
        c.fill = fill(bg)
        c.font = font(BLACK, sz=9)
        c.alignment = align("left", "center", wrap=True)
        c.border = border_thin("all")
    badge(ws2, row, 7, status, status_colors2.get(status, MGRAY),
          WHITE if status == "In progress" else BLACK)
    badge(ws2, row, 8, priority, priority_colors.get(priority, MGRAY), WHITE)

set_col_widths(ws2, [12, 9, 24, 44, 7, 24, 11, 8, 28])


# ═══════════════════════════════════════════════════════════════
# TAB 3 — DISTRIBUTION CHANNELS
# ═══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("3 · Distribution Channels")
ws3.sheet_view.showGridLines = False

merge_title(ws3, 1, 1, 8, "DISTRIBUTION CHANNELS  ·  Ranked by Audience Fit & Effort-to-Impact", BLACK, WHITE, 16)

header_row(ws3, 2, [
    "Priority", "Channel", "Audience Fit", "Est. Reach",
    "Effort", "When", "Action", "Status"
], NAVY, WHITE, height=22)

channels = [
    # P0
    ("P0", "Show HN (Hacker News)",
     "Developers building AI — exact audience",
     "50K–500K impressions if front page",
     "Low (1 post + 4h present)",
     "May 2026",
     "Post Mon/Tue morning ET. Title: 'Show HN: Regula — scan your AI project for EU AI Act risk in 10 seconds'. Link to GitHub repo. Be honest about precision.",
     "Not started"),
    ("P0", "artificialintelligenceact.eu",
     "EU AI Act practitioners — highest conversion",
     "150K+ users/month (self-reported)",
     "Med (outreach + article pitch)",
     "Jun 2026",
     "Contact site editors. Submit to tools directory. Offer guest article: 'Code-level EU AI Act risk detection — how it works'.",
     "Not started"),
    ("P0", "DEV.to article",
     "Developers — curious, searchable",
     "10K–100K reads (SEO + community)",
     "Med (research + write)",
     "May 2026",
     "Title: 'I scanned 5 open-source AI projects for EU AI Act compliance'. Run regula on LangChain/Ollama/AutoGPT. Show real output. Honest FP rate.",
     "Not started"),
    # P1
    ("P1", "EU Digital SME Alliance",
     "SMEs building AI in EU — budget-constrained",
     "Directory listing (unknown visitors)",
     "Low (form submission)",
     "Jun 2026",
     "Submit via their tool listing form. Emphasise: free, no account, 60 seconds, zero dependencies.",
     "Not started"),
    ("P1", "awesome-claude-code PR",
     "Claude Code users — overlapping toolchain",
     "Niche but targeted",
     "Low (1 PR)",
     "May 2026",
     "Submit PR. Regula has Claude Code hooks. Mention: regula-ignore works in AI-generated code.",
     "Not started"),
    ("P1", "GitHub Topics (#eu-ai-act)",
     "Devs searching compliance tools on GitHub",
     "Organic search traffic",
     "Done (topics set)",
     "Done",
     "Topics already set: eu-ai-act, compliance, ai-safety, static-analysis, python, cli.",
     "Done"),
    ("P1", "PyPI discoverability",
     "Python developers installing AI tools",
     "PyPI search + pip install",
     "Low (keywords in pyproject.toml)",
     "Apr 2026",
     "Add classifiers to pyproject.toml: 'Topic :: Security', 'Topic :: Software Development :: Quality Assurance'. Verify description renders on PyPI.",
     "Not started"),
    ("P1", "r/python + r/programming",
     "Python/developer community",
     "5K–50K views if popular",
     "High (participate first, 4 weeks)",
     "Nov 2026",
     "80/20 rule: participate in EU AI Act threads for 4 weeks. Then post: 'I scanned the top 50 PyPI AI packages for EU AI Act compliance — here's what I found'.",
     "Not started"),
    # P2
    ("P2", "LinkedIn content series",
     "CTOs, DPOs, AI engineers",
     "100–1,000 per post (cold start)",
     "Med (4 posts/month)",
     "Jul 2026",
     "4-post series: 'What the EU AI Act actually requires from your code, article by article'. Each post ends with a 'regula check .' call to action.",
     "Not started"),
    ("P2", "Product Hunt",
     "Broad tech audience — less targeted",
     "500–5,000 if top 10",
     "Med (prep + launch day)",
     "Jun 2026",
     "Second wave after HN. Create product page. Screenshots of CLI output. Schedule Tuesday launch.",
     "Not started"),
    ("P2", "IAPP community",
     "Privacy/compliance professionals",
     "Unknown — conservative audience",
     "High (community membership first)",
     "Jul 2026",
     "Attend IAPP forums. Contribute to AI Act compliance matrix discussion. Frame Regula as triage tool for DPOs, not a replacement for legal review.",
     "Not started"),
    ("P2", "MCP server directories",
     "Claude/Cursor users",
     "Niche and growing",
     "Low (listing submission)",
     "May 2026",
     "List the Regula MCP server in Claude/Cursor tool directories. Good for developer mindshare.",
     "Not started"),
    # P3
    ("P3", "AI Act Service Desk (EC)",
     "EU startups/SMEs seeking tools",
     "Low visibility — uncertain",
     "Low",
     "Aug 2026",
     "Submit around the August deadline when traffic to official resources peaks.",
     "Not started"),
    ("P3", "YouTube tutorial",
     "Visual learners, international reach",
     "Long-tail SEO value",
     "High (record, edit, upload)",
     "Sep 2026",
     "5-minute tutorial: 'Run regula check on your AI project'. Show prohibited → high-risk → compliant flow. Real code, real output.",
     "Not started"),
]

ch_priority_colors = {"P0": ("FFE8E8", RED_), "P1": ("FFF3E0", ORANGE), "P2": ("FFFDE7", AMBER), "P3": ("E8F5E9", GREEN)}

for i, (pr, ch, fit, reach, effort, when, action, status) in enumerate(channels):
    row = 3 + i
    ws3.row_dimensions[row].height = 65
    bg, _ = ch_priority_colors.get(pr, (WHITE, BLACK))
    for col, val in enumerate([None, ch, fit, reach, effort, when, action, status], 1):
        if col == 1:
            badge(ws3, row, 1, pr, ch_priority_colors[pr][1], WHITE)
            continue
        c = ws3.cell(row=row, column=col, value=val)
        c.fill = fill(bg)
        c.font = font(BLACK, sz=9)
        c.alignment = align("left", "center", wrap=True)
        c.border = border_thin("all")
    # status badge override
    sc = GREEN if status == "Done" else (AMBER if status == "In progress" else MGRAY)
    badge(ws3, row, 8, status, sc, WHITE if status == "Done" else BLACK)

set_col_widths(ws3, [8, 22, 22, 22, 18, 10, 50, 12])


# ═══════════════════════════════════════════════════════════════
# TAB 4 — COMPETITIVE LANDSCAPE
# ═══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("4 · Competitive Landscape")
ws4.sheet_view.showGridLines = False

merge_title(ws4, 1, 1, 9, "COMPETITIVE LANDSCAPE  ·  Verified Data, April 2026", BLACK, WHITE, 16)
merge_title(ws4, 2, 1, 9, "All star counts verified via GitHub API. All claims verified against public sources.", DGRAY, WHITE, 12)

header_row(ws4, 3, [
    "Tool", "Stars\n(verified)", "Type", "Approach", "CLI?", "CI/CD?",
    "Languages", "Regula Advantage", "Risk to Regula"
], NAVY, WHITE, height=30)

competitors = [
    ("Regula (ours)",    "0 ★",   "CLI / code scanner",       "Regex + AST + tree-sitter. 33 commands, 53 patterns, 11 frameworks, 8 languages. Generates Annex IV docs, evidence packs, remediation plans.",  "Yes", "Yes (action.yml)", "8 (Py/JS/TS/Java/Go/Rust/C/C++)", "Most commands (33), most languages (8), only tool with Annex IV doc generation + evidence pack + gap scoring. Free.", "Starting from 0. Late entrant. Systima overlaps on CLI+CI/CD."),
    ("EuConform",        "107 ★", "Browser app",              "Risk classification (Art.6/7) + bias eval (CrowS-Pairs). Offline-first, no cloud. PDF reports.",                       "No",  "No",               "N/A (form-based)",               "No CLI, no CI/CD. Cannot integrate into a dev workflow. No codebase scanning.", "107 stars vs 0 — strong first-mover advantage in OSS space"),
    ("Systima Comply",   "0 ★",   "CLI / GitHub Action",      "npm package + GitHub Action + TypeScript API. Scans codebase for EU AI Act risks. Domain-based severity. PDF reports. Call-chain tracing.", "Yes", "Yes (GH Action)",  "JS/TS (npm)",                    "Regula covers 8 languages vs JS/TS only. Regula has 33 commands and generates Annex IV docs; Systima focuses on scanning.", "Direct competitor — CLI + CI/CD + codebase scanning. Created 14 Mar 2026."),
    ("AgentGuard",       "10 ★",  "Runtime middleware",       "3-line Python import. Wraps LLM agents. Runtime policy enforcement, not code scanning.",                                "No",  "Yes (SDK)",        "Python (middleware)",             "Different use case: runtime vs scan-time. Complementary, not competitive.", "Could expand to code scanning"),
    ("EU AI Radar",      "?",     "Static web tool",          "5-question quiz. Maps to risk band. No code scanning.",                                                                 "No",  "No",               "N/A",                            "Trivial tool. Different depth entirely.",                                   "Very low — different depth"),
    ("mcp-eu-ai-act",    "2 ★",   "MCP server",               "ArkForge MCP scanner. Detects EU AI Act violations via MCP protocol.",                                                 "Via MCP", "Via MCP",     "MCP-compatible",                 "Very early. Our MCP server is a distribution channel, not a competitor.", "Could gain traction with Claude/Cursor users before us"),
    ("G0 (AgentBouncr)", "?",     "Agent control layer",      "Scan, test, monitor for AI agents. LangChain/CrewAI/AutoGen/RAG. HMAC audit chains. ConsentGate.",                    "No",  "Yes",              "Python",                         "Agent-focused vs code-focused. Different buyer.",                          "Could expand to code scanning"),
    ("KLA Digital",      "N/A",   "SaaS platform",            "AI governance SaaS. Cross-framework mapping. €299+/month.",                                                            "No",  "No",               "Platform",                        "We're free. They target enterprise compliance officers, not developers.",  "Credibility: established brand vs unknown"),
    ("OneTrust/Credo AI","N/A",   "Enterprise GRC",           "Est. $50K+/year for enterprise contracts. Risk management, documentation, lifecycle.",                                 "No",  "No",               "Platform",                        "We're free. Entirely different buyer. Not a real competitor at our stage.","Credibility comparison if enterprise asks why not use them"),
]

comp_row_bgs = [
    ("1A1A2E", "EBF5FB"),  # Regula — highlighted
]
for i in range(1, len(competitors)):
    comp_row_bgs.append(("", "LGRAY" if i % 2 == 0 else "WHITE"))

for i, (tool, stars, typ, approach, cli, cicd, langs, advantage, risk) in enumerate(competitors):
    row = 4 + i
    ws4.row_dimensions[row].height = 60
    bg = "EBF5FB" if i == 0 else (LGRAY if i % 2 == 0 else WHITE)
    fg = BLACK
    bold = i == 0
    for col, val in enumerate([tool, stars, typ, approach, cli, cicd, langs, advantage, risk], 1):
        c = ws4.cell(row=row, column=col, value=val)
        c.fill = fill(bg)
        c.font = font(fg, bold=bold, sz=9)
        c.alignment = align("left", "center", wrap=True)
        c.border = border_thin("all")
    # Highlight CLI/CI cells
    for col, val in [(5, cli), (6, cicd)]:
        c = ws4.cell(row=row, column=col)
        if val in ("Yes", "Yes (action.yml)", "Yes (SDK)", "Via MCP"):
            c.fill = fill("E8F8F0")
            c.font = font(GREEN, bold=True, sz=9)
        else:
            c.fill = fill("FFE8E8")
            c.font = font(RED_, bold=True, sz=9)

set_col_widths(ws4, [18, 10, 16, 45, 8, 12, 20, 38, 30])

# Honest summary
ws4.row_dimensions[13].height = 10
merge_title(ws4, 14, 1, 9, "HONEST ASSESSMENT", ACCENT, WHITE, 13)
ws4.row_dimensions[15].height = 70
ws4.merge_cells("A15:I15")
c = ws4.cell(15, 1,
    "Regula has the broadest feature set in the open-source space (33 commands, 8 languages, 11 frameworks, Annex IV doc generation). "
    "One direct competitor on CLI + CI/CD exists: Systima Comply (npm/GitHub Action, JS/TS only, created 14 Mar 2026, 0 stars). "
    "Regula's advantage over Systima is language breadth (8 vs 2) and depth (33 commands vs a scanner). EuConform leads on OSS mindshare (107 stars).\n\n"
    "The honest risk: EuConform has 107 stars and 4 months head start. Systima is a newer but direct competitor. "
    "Distribution — not product quality — is the critical path. "
    "Every week without a Show HN post or distribution action is a week competitors compound their lead.\n\n"
    "The window is August 2026. After the deadline passes, urgency drops and the market fragments further.")
c.fill = fill(LGRAY)
c.font = font(BLACK, sz=10)
c.alignment = align("left", "top", wrap=True)
c.border = border_thin("all")


# ═══════════════════════════════════════════════════════════════
# TAB 5 — QUARTERLY OKR TRACKER
# ═══════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("5 · OKR Tracker")
ws5.sheet_view.showGridLines = False

merge_title(ws5, 1, 1, 8, "QUARTERLY OKR TRACKER  ·  Update Weekly", BLACK, WHITE, 16)

quarters = [
    ("Q1 Apr–Jun 2026", "Pre-launch & launch", BLUE),
    ("Q2 Jul–Sep 2026", "Ecosystem & deadline window", TEAL),
    ("Q3 Oct–Dec 2026", "Growth & community", PURPLE),
    ("Q4 Jan–Mar 2027", "Sustainability & v2 planning", ACCENT),
]

q_okrs = [
    # Q1
    [
        ("GitHub stars",              "0",   "50",   "", "", ""),
        ("PyPI downloads (total)",    "0",   "200",  "", "", ""),
        ("Show HN submitted",         "No",  "Yes",  "", "", ""),
        ("DEV.to article published",  "No",  "Yes",  "", "", ""),
        ("Landing page verified live","No",  "Yes",  "", "", ""),
        ("awesome-claude-code PR",    "No",  "Yes",  "", "", ""),
    ],
    # Q2
    [
        ("GitHub stars",              "?",   "200",  "", "", ""),
        ("PyPI monthly downloads",    "?",   "500",  "", "", ""),
        ("artificialintelligenceact.eu listed", "No", "Yes", "", "", ""),
        ("LinkedIn posts published",  "0",   "4",    "", "", ""),
        ("Real-world eval published", "No",  "Yes",  "", "", ""),
        ("Aug 2 blog post live",      "No",  "Yes",  "", "", ""),
    ],
    # Q3
    [
        ("GitHub stars",              "?",   "500",  "", "", ""),
        ("PyPI monthly downloads",    "?",   "1,000","", "", ""),
        ("User-filed GitHub issues",  "0",   "10",   "", "", ""),
        ("v1.6.0 shipped",            "No",  "Yes",  "", "", ""),
        ("r/python post submitted",   "No",  "Yes",  "", "", ""),
        ("False positive rate (OSS)", "N/A", "< 5%", "", "", ""),
    ],
    # Q4
    [
        ("GitHub stars",              "?",   "800",  "", "", ""),
        ("PyPI monthly downloads",    "?",   "2,000","", "", ""),
        ("GitHub Sponsors live",      "No",  "Yes",  "", "", ""),
        ("Enterprise enquiries",      "0",   "3",    "", "", ""),
        ("Year-1 retro published",    "No",  "Yes",  "", "", ""),
        ("v2.0 roadmap defined",      "No",  "Yes",  "", "", ""),
    ],
]

current_row = 3
for (q_label, q_theme, q_color), q_data in zip(quarters, q_okrs):
    merge_title(ws5, current_row, 1, 8, f"{q_label}  ·  Theme: {q_theme}", q_color, WHITE, 13)
    current_row += 1
    header_row(ws5, current_row, [
        "Key Result", "Baseline", "Target",
        "Week 1", "Week 2", "Week 3", "Week 4", "RAG Status"
    ], NAVY, WHITE, height=22)
    current_row += 1
    for row_data in q_data:
        metric, baseline, target = row_data[0], row_data[1], row_data[2]
        w1, w2, w3, w4 = "", "", "", ""
        ws5.row_dimensions[current_row].height = 22
        data_row(ws5, current_row, [metric, baseline, target, w1, w2, w3, w4, ""], bg=LGRAY if current_row % 2 == 0 else WHITE)
        # RAG cell for user to fill
        c = ws5.cell(current_row, 8)
        c.fill = fill(MGRAY)
        c.font = font(DGRAY, sz=9, italic=True)
        c.value = "R / A / G"
        c.alignment = align("center", "center")
        current_row += 1
    current_row += 1  # spacer

set_col_widths(ws5, [36, 12, 12, 10, 10, 10, 10, 12])

# Instructions
merge_title(ws5, current_row, 1, 8, "HOW TO USE THIS TRACKER", DGRAY, WHITE, 12)
current_row += 1
instructions = [
    "Update the Week columns each Friday. Enter actual numbers, not estimates.",
    "RAG = Red (off track, needs action) / Amber (at risk, monitor) / Green (on track).",
    "A metric that stays Red for 2+ weeks needs a plan change, not a colour change.",
    "Baseline = value at start of quarter. Target = realistic stretch, based on market data above.",
    "Do not update baselines mid-quarter — that hides regression.",
]
for inst in instructions:
    ws5.row_dimensions[current_row].height = 20
    ws5.merge_cells(f"A{current_row}:H{current_row}")
    c = ws5.cell(current_row, 1, f"• {inst}")
    c.fill = fill(LGRAY)
    c.font = font(DGRAY, sz=10, italic=True)
    c.alignment = align("left", "center")
    current_row += 1


# ═══════════════════════════════════════════════════════════════
# TAB 6 — FIRST 8 WEEKS SPRINT PLAN
# ═══════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("6 · First 8 Weeks")
ws6.sheet_view.showGridLines = False

merge_title(ws6, 1, 1, 6, "FIRST 8 WEEKS — SPRINT PLAN  ·  April 3 – May 29, 2026", BLACK, WHITE, 16)
merge_title(ws6, 2, 1, 6, "These 8 weeks determine whether the Show HN lands before or after the August deadline crowd", ACCENT, WHITE, 12)

header_row(ws6, 3, ["Week", "Dates", "Focus", "Must-Do Tasks", "Done?", "Notes"], NAVY, WHITE, height=22)

weeks = [
    ("Wk 1", "Apr 3–9",  "Pre-launch: repo & landing",
     "□ Verify getregula.com is live and loads correctly\n"
     "□ Test email capture form end-to-end\n"
     "□ Add README hero screenshot or GIF of 'regula check' output\n"
     "□ Add PyPI classifiers to pyproject.toml (Topic :: Security etc.)\n"
     "□ Verify PyPI page renders description correctly",
     "", "Blocking: landing page must be verified before HN"),
    ("Wk 2", "Apr 10–16", "Pre-launch: DEV.to research",
     "□ Run 'regula check .' on LangChain (pip install langchain)\n"
     "□ Run on Ollama Python client\n"
     "□ Run on PrivateGPT or similar\n"
     "□ Document findings honestly (what it caught, what it missed)\n"
     "□ Draft DEV.to article outline",
     "", "Need 3-5 real repos with honest output before publishing"),
    ("Wk 3", "Apr 17–23", "DEV.to article: write & publish",
     "□ Write full article: 'I scanned 5 open-source AI repos for EU AI Act compliance'\n"
     "□ Include real terminal output (screenshots or code blocks)\n"
     "□ Be honest about false positives\n"
     "□ Publish on DEV.to\n"
     "□ Share on LinkedIn",
     "", "Article is evidence. Show the tool working, not just describe it."),
    ("Wk 4", "Apr 24–30", "Show HN prep",
     "□ Read 10 recent successful Show HN posts — note what made them land\n"
     "□ Finalise HN post title (from SHOW_HN_DRAFT.md)\n"
     "□ Prepare answers to all 6 predictable questions in draft\n"
     "□ Choose a Tuesday or Wednesday morning to post\n"
     "□ Block 4 hours in calendar for that day",
     "", "Timing matters. Clear your calendar for launch day."),
    ("Wk 5", "May 1–7",  "Show HN LAUNCH",
     "□ POST Show HN on chosen day, 8–10am ET\n"
     "□ Be present for minimum 4 hours\n"
     "□ Answer every technical question honestly and promptly\n"
     "□ Do NOT be defensive about false positives — acknowledge them\n"
     "□ Thank people who try it and report results",
     "", "This is the most important week. Clear everything else."),
    ("Wk 6", "May 8–14",  "Post-HN: follow-up & ecosystem",
     "□ Review all HN comments for product feedback\n"
     "□ Submit to awesome-claude-code\n"
     "□ Submit to MCP server directories\n"
     "□ Respond to any emails/issues from HN traffic\n"
     "□ Note which questions came up most — update FAQ",
     "", "HN traffic is short-lived. Act on feedback immediately."),
    ("Wk 7", "May 15–21", "Product Hunt prep",
     "□ Create PH product page\n"
     "□ Write tagline and description\n"
     "□ Take clean screenshots of CLI output\n"
     "□ Line up 5 people to upvote on launch day\n"
     "□ Schedule for Tuesday launch",
     "", "PH is second wave — don't launch same week as HN"),
    ("Wk 8", "May 22–29", "EU ecosystem outreach begins",
     "□ Contact artificialintelligenceact.eu editors\n"
     "□ Submit to EU Digital SME Alliance directory\n"
     "□ Write first LinkedIn post in the 4-post series\n"
     "□ Review PyPI download stats — set Q2 baseline\n"
     "□ Review GitHub stars — assess HN impact honestly",
     "", "If HN didn't land: diagnose why before next attempt"),
]

for i, (wk, dates, focus, tasks, done, notes) in enumerate(weeks):
    row = 4 + i
    ws6.row_dimensions[row].height = 90
    bg = LGRAY if i % 2 == 0 else WHITE
    for col, val in enumerate([wk, dates, focus, tasks, done, notes], 1):
        c = ws6.cell(row=row, column=col, value=val)
        c.fill = fill(bg)
        c.font = font(BLACK, sz=9, bold=(col in (1, 3)))
        c.alignment = align("left", "center", wrap=True)
        c.border = border_thin("all")
    # Done checkbox cell
    c = ws6.cell(row=row, column=5)
    c.fill = fill(MGRAY)
    c.font = font(DGRAY, sz=9, italic=True)
    c.value = "□ Not done"
    c.alignment = align("center", "center")

set_col_widths(ws6, [7, 11, 18, 60, 12, 32])


# ─── Save ───────────────────────────────────────────────────────
out_path = "docs/regula_gtm_plan.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
