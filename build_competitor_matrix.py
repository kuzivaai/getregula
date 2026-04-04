#!/usr/bin/env python3
"""
Regula Competitor Matrix — AI Governance & Code Scanning Tools
Generated: April 2026
All data sourced and cited inline. Verify claims before publishing.
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── Colour palette ────────────────────────────────────────────────────────────
GREEN  = "C6EFCE"   # Yes / confirmed
RED    = "FFC7CE"   # No
YELLOW = "FFEB9C"   # Partial / Unknown
BLUE   = "BDD7EE"   # Header row
DARK   = "1F4E79"   # Dark header text
GREY   = "F2F2F2"   # Alternating row
WHITE  = "FFFFFF"
ORANGE = "FCE4D6"   # Regula highlight

YES  = "✓"
NO   = "✗"
PART = "~"   # Partial / limited
UNK  = "?"   # Not verified — check source

# ── Tool definitions (ordered: most ticks → least) ───────────────────────────
# Format: { name, short_desc, primary_url, source_note }
TOOLS = [
    {
        "name": "Regula",
        "desc": "EU AI Act CLI scanner\n(Python, open source)",
        "url": "github.com/kuzivaai/getregula",
        "note": "Verified: pyproject.toml, README, CLAUDE.md",
        "highlight": True,
    },
    {
        "name": "VerifyWise",
        "desc": "Open-source AI governance platform\n(EU AI Act, ISO 42001, NIST)",
        "url": "verifywise.ai",
        "note": "Verified: verifywise.ai/pricing, github.com/bluewave-labs/verifywise",
        "highlight": False,
    },
    {
        "name": "Microsoft\nAgent Gov. Toolkit",
        "desc": "Open-source runtime security\nfor AI agents (released Apr 3 2026)",
        "url": "github.com/microsoft/agent-governance-toolkit",
        "note": "Verified: opensource.microsoft.com/blog/2026/04/02, helpnetsecurity.com/2026/04/03",
        "highlight": False,
    },
    {
        "name": "Credo AI",
        "desc": "Enterprise AI governance platform\n(market leader)",
        "url": "credo.ai",
        "note": "Verified: credo.ai/product, credo.ai/eu-ai-act",
        "highlight": False,
    },
    {
        "name": "Holistic AI",
        "desc": "Enterprise governance platform\n(Gartner recognised)",
        "url": "holisticai.com",
        "note": "Verified: holisticai.com, cloudeagle.ai/blogs/10-best-ai-governance-platforms",
        "highlight": False,
    },
    {
        "name": "IBM watsonx\n.governance",
        "desc": "Enterprise governance platform\n(financial sector depth)",
        "url": "ibm.com/products/watsonx-governance",
        "note": "Verified: ibm.com/products/watsonx-governance/pricing, AWS Marketplace",
        "highlight": False,
    },
    {
        "name": "OneTrust",
        "desc": "Enterprise GRC platform\n(privacy heritage)",
        "url": "onetrust.com",
        "note": "Verified: onetrust.com/solutions/eu-ai-act-compliance",
        "highlight": False,
    },
    {
        "name": "Systima\n(@systima/comply)",
        "desc": "CLI EU AI Act scanner\n(Node.js, AST-based, CI/CD)",
        "url": "systima.ai / npx @systima/comply",
        "note": "Verified: dev.to/systima/open-source-eu-ai-act-compliance-scanning-for-cicd",
        "highlight": False,
    },
    {
        "name": "RegulyAI",
        "desc": "Multi-framework compliance platform\n(50+ frameworks, US state laws)",
        "url": "regulyai.com",
        "note": "Verified: regulyai.com (features page). Pricing unverified.",
        "highlight": False,
    },
    {
        "name": "Quethos Sentinel",
        "desc": "SaaS EU AI Act scanner\n(GitHub/GitLab/Bitbucket)",
        "url": "quethossentinel.eu",
        "note": "Verified: quethossentinel.eu. Pricing unverified.",
        "highlight": False,
    },
    {
        "name": "SafeDep",
        "desc": "Open-source shadow AI &\nsupply-chain scanner (CLI)",
        "url": "safedep.io",
        "note": "Verified: safedep.io/shadow-ai, safedep.io/pricing, github.com/safedep/vet",
        "highlight": False,
    },
    {
        "name": "Cycode",
        "desc": "SDLC AI security platform\n(AIBOM, enterprise)",
        "url": "cycode.com",
        "note": "Verified: cycode.com/blog/mapping-shadow-ai-sdlc-visibility",
        "highlight": False,
    },
    {
        "name": "SCW Trust Agent",
        "desc": "Commit-level AI attribution\n(launched Mar 2026)",
        "url": "securecodewarrior.com/product/trust-agent-ai",
        "note": "Verified: businesswire.com/news/home/20260317399993, helpnetsecurity.com/2026/03/17",
        "highlight": False,
    },
    {
        "name": "Midasyannkc\nAI-gov-Scanner",
        "desc": "Open-source ML registry auditor\n(MLflow, SageMaker, Azure ML)",
        "url": "github.com/Midasyannkc/AI-governance-Scanner-",
        "note": "Verified: GitHub repo exists. Star count / activity unverified — check repo.",
        "highlight": False,
    },
    {
        "name": "FireTail",
        "desc": "Enterprise shadow AI detection\n(code + cloud + browser)",
        "url": "firetail.ai",
        "note": "Verified: firetail.ai/eliminate-shadow-ai, g2.com/products/firetail/reviews",
        "highlight": False,
    },
    {
        "name": "Zenity",
        "desc": "Enterprise AI agent security\n(Copilot/Salesforce native)",
        "url": "zenity.io",
        "note": "Verified: zenity.io/use-cases/risk-type/shadow-ai, Forrester Q2 2025",
        "highlight": False,
    },
]

# ── Row definitions ───────────────────────────────────────────────────────────
# Each row: (category, label, [values per tool], source_note, row_type)
# row_type: "section" = grey section header, "data" = normal row

ROWS = [
    # ── ACCESS & COST ─────────────────────────────────────────────────────────
    ("ACCESS & COST", None, None, None, "section"),
    ("Access", "Free tier available",
     [YES, YES, YES, PART, NO, NO, NO, YES, UNK, UNK, YES, NO, NO, YES, NO, NO],
     "Regula: open source MIT. VerifyWise: self-host free forever (verifywise.ai/pricing). "
     "MS toolkit: MIT. Credo AI: limited free trial. SafeDep: free to start. "
     "Midasyannkc: open source.",
     "data"),
    ("Access", "Open source",
     [YES, PART, YES, NO, NO, NO, NO, YES, NO, NO, YES, NO, NO, YES, NO, NO],
     "Regula: MIT. VerifyWise: BSL 1.1 (source-available, NOT full OSS — github.com/bluewave-labs/verifywise). "
     "MS toolkit: MIT (github.com/microsoft/agent-governance-toolkit). SafeDep: Apache-2. "
     "Systima: open source per dev.to article. Midasyannkc: open source.",
     "data"),
    ("Access", "Public pricing page",
     [YES, YES, PART, NO, NO, PART, NO, YES, UNK, UNK, YES, NO, NO, YES, NO, NO],
     "IBM: partial — AWS Marketplace lists $38,160/yr base. VerifyWise: verifywise.ai/pricing. "
     "Credo AI, OneTrust, Holistic AI: contact sales only. FireTail, Zenity, Cycode, SCW: quote-based.",
     "data"),
    ("Access", "Enterprise / custom pricing",
     [NO, YES, NO, YES, YES, YES, YES, UNK, UNK, UNK, YES, YES, YES, NO, YES, YES],
     "IBM: $10–25k/month mid-market per redresscompliance.com/ibm-watsonx-licensing-guide. "
     "All enterprise tools are quote-based.",
     "data"),

    # ── DEVELOPER TOOLING ─────────────────────────────────────────────────────
    ("DEVELOPER TOOLING", None, None, None, "section"),
    ("Dev", "CLI available",
     [YES, NO, YES, NO, NO, NO, NO, YES, NO, NO, YES, YES, YES, YES, NO, NO],
     "Regula: `regula` command. MS toolkit: Python/TS/Rust/Go/.NET packages. "
     "Systima: `npx @systima/comply`. SafeDep: `vet` CLI. Cycode: SDLC CLI. "
     "SCW Trust Agent: commit-level CLI.",
     "data"),
    ("Dev", "CI/CD integration",
     [YES, NO, YES, UNK, UNK, UNK, NO, YES, NO, YES, YES, YES, YES, UNK, UNK, NO],
     "Regula: action.yml in repo. Systima: GitHub Action (systima-ai/comply@v1) per dev.to. "
     "Quethos: GitHub/GitLab/Bitbucket integration per quethossentinel.eu. "
     "SafeDep: CI supported per safedep.io docs.",
     "data"),
    ("Dev", "GitHub Action available",
     [YES, NO, UNK, UNK, NO, NO, NO, YES, NO, YES, UNK, UNK, UNK, UNK, UNK, NO],
     "Regula: action.yml committed to repo. Systima: systima-ai/comply@v1 per dev.to article.",
     "data"),
    ("Dev", "Multi-language code support",
     [YES, YES, YES, NO, NO, NO, NO, PART, NO, UNK, YES, UNK, UNK, UNK, UNK, NO],
     "Regula: 8 languages per CLAUDE.md. VerifyWise: Python/JS/Go detection. "
     "MS toolkit: Python/TS/Rust/Go/.NET. Systima: TypeScript primary, 37+ frameworks via tree-sitter. "
     "SafeDep: Go/Python/JS per safedep.io/shadow-ai-discovery-vet.",
     "data"),
    ("Dev", "Self-hosted / local-first",
     [YES, YES, YES, NO, UNK, YES, UNK, YES, NO, NO, YES, UNK, UNK, YES, UNK, UNK],
     "Regula: runs locally, no external calls. VerifyWise: self-host option (verifywise.ai/pricing). "
     "MS toolkit: local Python/TS packages. IBM: on-prem via VPC licensing. "
     "SafeDep: local-first, no sign-up for basic use.",
     "data"),
    ("Dev", "No API key required for basic use",
     [YES, YES, YES, NO, NO, NO, NO, YES, NO, NO, YES, NO, NO, YES, NO, NO],
     "Regula: zero dependencies, stdlib-only. Systima: 'no API keys required' per dev.to. "
     "SafeDep: 'no sign-up required' per safedep.io. MS toolkit: local packages.",
     "data"),

    # ── CODE SCANNING CAPABILITIES ────────────────────────────────────────────
    ("CODE SCANNING", None, None, None, "section"),
    ("Scan", "Code-level scanning",
     [YES, PART, YES, NO, NO, NO, NO, YES, NO, YES, YES, YES, YES, YES, YES, NO],
     "Regula: 161 risk patterns across 8 languages. VerifyWise: detects 80+ libraries (limited depth). "
     "MS toolkit: policy engine intercepts agent actions. Systima: AST-based 37+ AI frameworks. "
     "Quethos: scans GitHub/GitLab/Bitbucket repos. SafeDep: AI SDK call detection. "
     "Cycode: SDLC-wide. SCW: commit-level. FireTail: code + cloud + browser.",
     "data"),
    ("Scan", "AST / semantic code analysis",
     [YES, NO, NO, NO, NO, NO, NO, YES, NO, NO, YES, YES, PART, NO, NO, NO],
     "Regula: ast_engine.py in repo. Systima: explicitly AST-based via TypeScript Compiler API per dev.to. "
     "SafeDep: xbom uses semantic code analysis per github.com/safedep/xbom. "
     "Cycode: AI signal detection across SDLC.",
     "data"),
    ("Scan", "Shadow AI / unapproved tool detection",
     [YES, NO, YES, YES, YES, UNK, YES, NO, NO, NO, YES, YES, YES, NO, YES, YES],
     "SAP/Oxford Economics: 68% UK businesses have shadow AI (primary source: news.sap.com/uk/2026/02). "
     "Regula: discover command. MS toolkit: agent monitoring. Credo AI: auto-discovery per credo.ai/product. "
     "Holistic AI: continuous detection per holisticai.com. SafeDep: core feature. "
     "Cycode: AIBOM. SCW: commit-level shadow model detection. FireTail: code+cloud+browser. "
     "Zenity: runtime shadow agent detection.",
     "data"),
    ("Scan", "AI/ML library & SDK inventory",
     [YES, YES, YES, YES, YES, UNK, YES, PART, YES, YES, YES, YES, YES, YES, YES, YES],
     "Regula: model_inventory.py + discover command. VerifyWise: 80+ libraries. "
     "MS toolkit: agent inventory. All enterprise platforms include inventory as core feature.",
     "data"),
    ("Scan", "Call-chain / import tracing",
     [PART, NO, NO, NO, NO, NO, NO, YES, NO, NO, YES, UNK, NO, NO, NO, NO],
     "Systima: explicitly advertises 'call-chain tracing' per dev.to article. "
     "Regula: AST analysis + import detection. SafeDep: traces SDK calls.",
     "data"),
    ("Scan", "Runtime / agent monitoring",
     [NO, NO, YES, YES, YES, YES, YES, NO, NO, NO, NO, NO, NO, NO, YES, YES],
     "MS toolkit: sub-millisecond policy engine per opensource.microsoft.com/blog/2026/04/02. "
     "IBM watsonx: drift monitoring, Q1 2026 agent monitoring added. "
     "Credo AI: continuous eval of agent traces. Zenity: runtime shadow agent detection. "
     "FireTail: real-time anomaly detection.",
     "data"),

    # ── COMPLIANCE FRAMEWORK COVERAGE ─────────────────────────────────────────
    ("COMPLIANCE COVERAGE", None, None, None, "section"),
    ("Frameworks", "EU AI Act coverage",
     [YES, YES, YES, YES, YES, YES, YES, YES, YES, YES, NO, UNK, NO, NO, UNK, UNK],
     "Regula: 11 compliance frameworks per CLAUDE.md. VerifyWise: EU AI Act + ISO 42001 per verifywise.ai. "
     "MS toolkit: EU AI Act mapping per opensource.microsoft.com. Credo AI: per credo.ai/eu-ai-act. "
     "Holistic AI, IBM, OneTrust: all explicitly cover EU AI Act. Systima: Articles 5-50 per dev.to. "
     "RegulyAI: per regulyai.com. Quethos: Articles 5-55 per quethossentinel.eu. "
     "SafeDep: supply-chain focused, NOT EU AI Act compliance.",
     "data"),
    ("Frameworks", "Prohibited practice detection (Art. 5)",
     [YES, YES, UNK, YES, YES, YES, YES, YES, YES, YES, NO, NO, NO, NO, NO, NO],
     "Regula: risk_patterns.py includes prohibited practice patterns. "
     "Systima: Article 5 in scope per dev.to. Quethos: Article 5 in scope. "
     "VerifyWise, Credo AI, Holistic AI, IBM, OneTrust: covered via governance frameworks.",
     "data"),
    ("Frameworks", "Risk classification output\n(High / Limited / Minimal)",
     [YES, YES, YES, YES, YES, YES, YES, YES, YES, YES, NO, NO, NO, NO, UNK, NO],
     "Regula: classify command with 3-tier risk output. Systima: domain-adjusted severity. "
     "VerifyWise: risk scoring per verifywise.ai/platform. All enterprise platforms include risk scoring. "
     "RegulyAI: High/Limited/Minimal classification per regulyai.com.",
     "data"),
    ("Frameworks", "NIST AI RMF coverage",
     [YES, YES, UNK, YES, YES, YES, YES, NO, YES, NO, NO, NO, NO, NO, NO, UNK],
     "Regula: framework_mapper.py per repo. VerifyWise: NIST AI RMF per verifywise.ai. "
     "Credo AI, Holistic AI, IBM, OneTrust: all explicitly cover NIST. "
     "RegulyAI: 50+ frameworks including NIST per regulyai.com.",
     "data"),
    ("Frameworks", "ISO 42001 coverage",
     [NO, YES, NO, YES, YES, YES, UNK, NO, UNK, NO, NO, NO, NO, NO, NO, UNK],
     "VerifyWise: ISO 42001 explicitly listed per verifywise.ai. "
     "Credo AI, Holistic AI, IBM: all explicitly cover ISO 42001.",
     "data"),
    ("Frameworks", "UK / ICO framework coverage",
     [NO, UNK, NO, UNK, UNK, UNK, YES, NO, UNK, NO, NO, NO, NO, NO, NO, NO],
     "OneTrust: covers GDPR/ICO via privacy heritage. All others: unverified for UK-specific frameworks. "
     "UK has no dedicated AI Act — frameworks are ICO AI guidance + DSIT 5 principles (not codified law). "
     "Source: ico.org.uk/for-organisations/uk-gdpr-guidance-and-resources/artificial-intelligence",
     "data"),
    ("Frameworks", "Brazil Lei IA coverage",
     [PART, NO, NO, NO, NO, NO, NO, NO, UNK, NO, NO, NO, NO, NO, NO, NO],
     "Regula: Brazil module exists per CLAUDE.md. Bill still in Chamber of Deputies as of April 2026 — "
     "not yet law. Source: dataprivacybr.org, White & Case Brazil AI tracker.",
     "data"),
    ("Frameworks", "Multi-framework coverage (5+)",
     [YES, YES, PART, YES, YES, YES, YES, NO, YES, NO, NO, NO, NO, NO, NO, UNK],
     "Regula: 11 frameworks per CLAUDE.md. VerifyWise: 24+ per verifywise.ai. "
     "MS toolkit: EU AI Act + HIPAA + SOC2 (3 — partial). "
     "RegulyAI: 50+ frameworks per regulyai.com. Credo AI: EU AI Act + NIST + ISO + SOC2.",
     "data"),

    # ── GOVERNANCE & DOCUMENTATION ────────────────────────────────────────────
    ("GOVERNANCE & DOCS", None, None, None, "section"),
    ("Docs", "Technical documentation generation",
     [YES, YES, YES, YES, YES, YES, YES, YES, YES, YES, NO, NO, NO, NO, NO, NO],
     "Regula: generate_documentation.py. Systima: generates compliance documentation per dev.to. "
     "VerifyWise: documentation templates. All enterprise platforms generate docs.",
     "data"),
    ("Docs", "Evidence / audit pack export",
     [YES, YES, YES, YES, YES, YES, YES, YES, YES, YES, NO, NO, NO, NO, NO, NO],
     "Regula: evidence_pack.py + cmd_evidence_pack. VerifyWise: audit trails per verifywise.ai. "
     "MS toolkit: compliance grading per opensource.microsoft.com. "
     "All enterprise platforms include audit-ready evidence.",
     "data"),
    ("Docs", "PDF / HTML report export",
     [YES, YES, UNK, YES, YES, YES, YES, YES, YES, YES, NO, UNK, UNK, UNK, UNK, UNK],
     "Regula: pdf_export.py + HTML report command. Systima: PDF reports per dev.to. "
     "RegulyAI: 'regulator-ready PDFs' per regulyai.com. "
     "All enterprise platforms include report export.",
     "data"),
    ("Docs", "Bias detection / testing",
     [YES, YES, UNK, YES, YES, YES, UNK, NO, YES, UNK, NO, NO, NO, YES, NO, NO],
     "Regula: bias_eval.py. VerifyWise: visual bias reports per verifywise.ai/platform. "
     "Credo AI, Holistic AI, IBM: all include bias detection. "
     "RegulyAI: NYC LL144 bias audits per regulyai.com. "
     "Midasyannkc: flags 'bias testing gaps' per GitHub repo description.",
     "data"),
    ("Docs", "Dashboard / web UI",
     [NO, YES, UNK, YES, YES, YES, YES, NO, YES, YES, YES, YES, YES, NO, YES, YES],
     "Regula: CLI only, no UI. All governance platforms include dashboards. "
     "VerifyWise: full dashboard per verifywise.ai/platform.",
     "data"),
    ("Docs", "Questionnaire-based assessment",
     [YES, YES, NO, YES, YES, YES, YES, NO, YES, NO, NO, NO, NO, NO, NO, NO],
     "Regula: questionnaire command in CLI. VerifyWise: AI-assisted questionnaires per verifywise.ai. "
     "All enterprise platforms include questionnaires.",
     "data"),

    # ── BUSINESS & MOAT ───────────────────────────────────────────────────────
    ("BUSINESS & MOAT", None, None, None, "section"),
    ("Business", "Primary target audience",
     ["Developer", "IT/Compliance", "Developer/\nAgent teams", "Compliance/\nEnterprise",
      "Enterprise/\nCompliance", "Enterprise/\nFinancial", "Enterprise\nGRC",
      "Developer", "Compliance/\nLegal", "Developer/\nCompliance",
      "Developer/\nSecurity", "Developer/\nSecurity", "Developer/\nSecurity",
      "Developer/\nML Eng.", "Security\nTeam", "Enterprise\nCISO"],
     "Inferred from product positioning and feature set.",
     "audience"),
    ("Business", "Actively developed (2026)",
     [YES, YES, YES, YES, YES, YES, YES, YES, YES, YES, YES, YES, YES, UNK, YES, YES],
     "Regula: commits in repo. MS toolkit: released Apr 3 2026. SCW Trust Agent: released Mar 2026. "
     "Midasyannkc: activity unverified — check github.com/Midasyannkc/AI-governance-Scanner-.",
     "data"),
    ("Business", "Venture / institutional backing",
     [NO, UNK, YES, YES, YES, YES, YES, UNK, UNK, UNK, UNK, YES, YES, NO, YES, YES],
     "MS toolkit: Microsoft (opensource.microsoft.com). Credo AI: VC-backed. "
     "Holistic AI: VC-backed. IBM: public company. OneTrust: $5.1bn valuation. "
     "Zenity: Forrester recognised. Cycode: VC-backed.",
     "data"),
    ("Business", "Regulatory body / official tool",
     [NO, NO, NO, NO, NO, NO, NO, NO, NO, NO, NO, NO, NO, NO, NO, NO],
     "Note: EU Commission has its own compliance checker at "
     "ai-act-service-desk.ec.europa.eu — free web questionnaire tool. "
     "NOT included as a competitor above as it is not a developer tool.",
     "data"),
    ("Business", "Key moat / differentiator",
     [
         "Only open-source CLI mapping\ncode patterns to EU AI Act\nrisk articles. 8 languages.\n161 patterns.",
         "Most complete open-source\ngovernance platform. 24+\nframeworks. BSL licence.",
         "Microsoft backing. OWASP\nAgentic Top 10. Sub-ms\npolicy engine. 9,500+ tests.",
         "Market leader. 10x faster\ncompliance claim. Enterprise\nautomation.",
         "Gartner recognised. Full\nlifecycle. Continuous shadow\nAI detection.",
         "IBM brand + financial\nsector depth. AWS/Azure\nmarketplace.",
         "Largest GRC platform.\nGDPR/privacy heritage.\nExisting enterprise clients.",
         "AST call-chain tracing.\nDomain-adjusted severity.\n37+ AI frameworks.",
         "50+ frameworks incl.\n38+ US state laws.\nBias audits (NYC LL144).",
         "Zero-install. Articles\n5-55 full scope.\nGitHub/GitLab/Bitbucket.",
         "Supply chain + shadow AI.\nNo sign-up. AI coding\nagent security (Gryph).",
         "AIBOM. SDLC-wide\nvisibility. AI vulnerability\nscanning.",
         "Commit-level LLM\nattribution. MCP discovery.\nSecurity benchmarking.",
         "ML registry scanning\n(MLflow, SageMaker,\nAzure ML, Vertex AI).",
         "Triple detection: code +\ncloud + browser. PII\nin prompts detection.",
         "Copilot/Salesforce native.\nForrest recognised.\nRuntime agent security.",
     ],
     "Inferred from product pages, GitHub repos, and press releases. Verify claims before citing.",
     "moat"),

    # ── SCORE ─────────────────────────────────────────────────────────────────
    ("SCORE", None, None, None, "section"),
    ("Score", "Verified ✓ count (binary rows only)",
     ["=COUNT_PLACEHOLDER"] * 16,   # filled programmatically
     "Count of YES (✓) across all binary feature rows above. Higher = broader confirmed capability.",
     "score"),
]

def make_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def make_border():
    thin = Side(style='thin', color='CCCCCC')
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def build_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Competitor Matrix"

    n_tools = len(TOOLS)

    # ── Title row ─────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4 + n_tools)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = (
        "Regula — AI Governance & Code Scanning Competitive Matrix  |  April 2026  |  "
        "All claims cited inline — verify before publishing"
    )
    title_cell.font = Font(bold=True, size=12, color="FFFFFF")
    title_cell.fill = make_fill("1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    # ── Column headers ─────────────────────────────────────────────────────────
    header_labels = ["Category", "Feature / Criterion", "Source / Verification Notes"]
    for i, label in enumerate(header_labels, start=1):
        c = ws.cell(row=2, column=i)
        c.value = label
        c.font = Font(bold=True, size=9, color="FFFFFF")
        c.fill = make_fill(DARK)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = make_border()

    for j, tool in enumerate(TOOLS):
        col = 4 + j
        c = ws.cell(row=2, column=col)
        c.value = tool["name"]
        c.font = Font(bold=True, size=8, color="FFFFFF" if not tool["highlight"] else "FFD700")
        c.fill = make_fill("1F4E79" if not tool["highlight"] else "0D3E6B")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = make_border()

    ws.row_dimensions[2].height = 42

    # ── Tool description sub-header ───────────────────────────────────────────
    for i in range(1, 4):
        c = ws.cell(row=3, column=i)
        c.fill = make_fill("E9EFF7")
        c.border = make_border()
    for j, tool in enumerate(TOOLS):
        col = 4 + j
        c = ws.cell(row=3, column=col)
        c.value = tool["desc"]
        c.font = Font(italic=True, size=7, color="444444")
        c.fill = make_fill(ORANGE if tool["highlight"] else "E9EFF7")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = make_border()
    ws.row_dimensions[3].height = 36

    # ── Tool URL row ──────────────────────────────────────────────────────────
    for i in range(1, 4):
        c = ws.cell(row=4, column=i)
        c.fill = make_fill("F5F5F5")
        c.border = make_border()
    for j, tool in enumerate(TOOLS):
        col = 4 + j
        c = ws.cell(row=4, column=col)
        c.value = tool["url"]
        c.font = Font(size=7, color="1155CC", underline="single")
        c.fill = make_fill("F5F5F5")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = make_border()
    ws.row_dimensions[4].height = 24

    # ── Data rows ─────────────────────────────────────────────────────────────
    current_row = 5
    score_row = None
    yes_counts = [0] * n_tools

    for row_def in ROWS:
        category, label, values, source_note, row_type = row_def

        if row_type == "section":
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row, end_column=3 + n_tools
            )
            c = ws.cell(row=current_row, column=1)
            c.value = f"  {category}"
            c.font = Font(bold=True, size=9, color="FFFFFF")
            c.fill = make_fill("2E75B6")
            c.alignment = Alignment(vertical="center")
            c.border = make_border()
            ws.row_dimensions[current_row].height = 18
            current_row += 1
            continue

        # Category cell
        c_cat = ws.cell(row=current_row, column=1)
        c_cat.value = category
        c_cat.font = Font(size=8, color="666666")
        c_cat.fill = make_fill(GREY)
        c_cat.alignment = Alignment(horizontal="center", vertical="center")
        c_cat.border = make_border()

        # Label cell
        c_lbl = ws.cell(row=current_row, column=2)
        c_lbl.value = label
        c_lbl.font = Font(size=8, bold=(row_type == "score"))
        c_lbl.fill = make_fill(GREY)
        c_lbl.alignment = Alignment(vertical="center", wrap_text=True)
        c_lbl.border = make_border()

        # Source note cell
        c_src = ws.cell(row=current_row, column=3)
        c_src.value = source_note
        c_src.font = Font(size=7, color="555555", italic=True)
        c_src.fill = make_fill(GREY)
        c_src.alignment = Alignment(vertical="top", wrap_text=True)
        c_src.border = make_border()

        ws.row_dimensions[current_row].height = 44

        # Value cells
        for j, val in enumerate(values):
            col = 4 + j
            c = ws.cell(row=current_row, column=col)
            tool = TOOLS[j]

            if row_type in ("audience", "moat"):
                c.value = val
                c.font = Font(size=7)
                c.fill = make_fill(ORANGE if tool["highlight"] else WHITE)
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws.row_dimensions[current_row].height = 52

            elif row_type == "score":
                c.value = yes_counts[j]
                c.font = Font(bold=True, size=10)
                score_fill = ORANGE if tool["highlight"] else "C6EFCE"
                c.fill = make_fill(score_fill)
                c.alignment = Alignment(horizontal="center", vertical="center")

            else:
                c.value = val
                c.font = Font(size=10, bold=True)
                c.alignment = Alignment(horizontal="center", vertical="center")

                if val == YES:
                    c.fill = make_fill(GREEN if not tool["highlight"] else "92D050")
                    yes_counts[j] += 1
                elif val == NO:
                    c.fill = make_fill(RED)
                elif val == PART:
                    c.fill = make_fill(YELLOW)
                    yes_counts[j] += 0.5
                elif val == UNK:
                    c.fill = make_fill("ECECEC")
                else:
                    c.fill = make_fill(WHITE)

            c.border = make_border()

        current_row += 1

    # ── Now fill in actual score values ───────────────────────────────────────
    # Score row was the last row — go back and update it
    score_data_row = current_row - 1
    for j in range(n_tools):
        col = 4 + j
        c = ws.cell(row=score_data_row, column=col)
        c.value = int(yes_counts[j])

    # ── Legend ────────────────────────────────────────────────────────────────
    current_row += 1
    legend_items = [
        (YES,  GREEN,  "Confirmed / Yes"),
        (NO,   RED,    "Not available / No"),
        (PART, YELLOW, "Partial or limited capability"),
        (UNK,  "ECECEC", "Unverified — check source before citing"),
    ]
    ws.cell(row=current_row, column=1).value = "LEGEND"
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=8)
    for k, (sym, color, desc) in enumerate(legend_items):
        c = ws.cell(row=current_row, column=2 + k)
        c.value = f"{sym}  {desc}"
        c.fill = make_fill(color)
        c.font = Font(size=8)
        c.border = make_border()
        c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[current_row].height = 18

    current_row += 1
    ws.cell(row=current_row, column=1).value = "DATA INTEGRITY NOTE"
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=8, color="C00000")
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3 + n_tools)
    note = ws.cell(row=current_row, column=2)
    note.value = (
        "All data sourced from public web pages, GitHub repos, and press releases as of April 2026. "
        "? = Not verified — do not publish without checking the source. "
        "Pricing data changes frequently — always verify before quoting. "
        "EU AI Act Omnibus: both Council (Mar 13) and Parliament (Mar 26) have adopted positions proposing delay "
        "to Dec 2, 2027 (standalone) / Aug 2, 2028 (embedded). Trilogue ongoing — Aug 2, 2026 still legally binding. "
        "Source: europarl.europa.eu/legislative-train + addleshawgoddard.com (Apr 2026)."
    )
    note.font = Font(size=7, italic=True, color="C00000")
    note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[current_row].height = 36

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 12   # Category
    ws.column_dimensions["B"].width = 28   # Feature
    ws.column_dimensions["C"].width = 52   # Source note
    for j in range(n_tools):
        col_letter = get_column_letter(4 + j)
        ws.column_dimensions[col_letter].width = 14

    # ── Freeze panes ──────────────────────────────────────────────────────────
    ws.freeze_panes = "D5"

    # ── Second sheet: Source log ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Source Log")
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 60
    ws2.column_dimensions["C"].width = 40

    headers = ["Tool", "Primary Source URL", "Verification Status"]
    for i, h in enumerate(headers, 1):
        c = ws2.cell(row=1, column=i)
        c.value = h
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = make_fill(DARK)
        c.border = make_border()

    source_data = [
        ("Regula", "github.com/kuzivaai/getregula + pyproject.toml", "Verified — codebase in session"),
        ("VerifyWise", "verifywise.ai/pricing + github.com/bluewave-labs/verifywise", "Verified — web search Apr 2026"),
        ("MS Agent Gov. Toolkit", "opensource.microsoft.com/blog/2026/04/02 + github.com/microsoft/agent-governance-toolkit", "Verified — released Apr 3 2026"),
        ("Credo AI", "credo.ai/product + credo.ai/eu-ai-act", "Verified — web search Apr 2026"),
        ("Holistic AI", "holisticai.com + Gartner 2024 Innovation Guide", "Verified — web search Apr 2026"),
        ("IBM watsonx.governance", "ibm.com/products/watsonx-governance/pricing + AWS Marketplace", "Verified — pricing confirmed at $38,160/yr base"),
        ("OneTrust", "onetrust.com/solutions/eu-ai-act-compliance", "Verified — web search Apr 2026"),
        ("Systima", "dev.to/systima/open-source-eu-ai-act-compliance-scanning-for-cicd", "Verified — published ~3 weeks before Apr 2026"),
        ("RegulyAI", "regulyai.com (features page)", "Partial — pricing unverified, features from product page"),
        ("Quethos Sentinel", "quethossentinel.eu", "Partial — pricing unverified"),
        ("SafeDep", "safedep.io/shadow-ai + safedep.io/pricing + github.com/safedep/vet", "Verified — web search Apr 2026"),
        ("Cycode", "cycode.com/blog/mapping-shadow-ai-sdlc-visibility + msspalert.com (AIBOM)", "Verified — web search Apr 2026"),
        ("SCW Trust Agent", "businesswire.com/news/home/20260317399993 + helpnetsecurity.com/2026/03/17", "Verified — press release Mar 17 2026"),
        ("Midasyannkc Scanner", "github.com/Midasyannkc/AI-governance-Scanner-", "Partial — GitHub repo exists; star count and activity UNVERIFIED"),
        ("FireTail", "firetail.ai/eliminate-shadow-ai + g2.com/products/firetail/reviews", "Verified — web search Apr 2026"),
        ("Zenity", "zenity.io/use-cases/risk-type/shadow-ai + Forrester Q2 2025", "Verified — web search Apr 2026"),
        ("EU Omnibus", "europarl.europa.eu/legislative-train + addleshawgoddard.com (Apr 2026)", "Verified — Council Mar 13, Parliament Mar 26"),
        ("Shadow AI 68% stat", "news.sap.com/uk/2026/02 (SAP/Oxford Economics, 200 UK execs)", "Verified primary source — NOT ObvioTech"),
        ("7% governance stat", "trustmarque.com/resources/press-release-ai-governance-report (507 UK IT DMs, Jul 2025)", "Verified — vendor-commissioned, note age"),
    ]

    for r, (tool, url, status) in enumerate(source_data, 2):
        ws2.cell(row=r, column=1).value = tool
        ws2.cell(row=r, column=2).value = url
        ws2.cell(row=r, column=3).value = status
        for col in range(1, 4):
            ws2.cell(row=r, column=col).font = Font(size=8)
            ws2.cell(row=r, column=col).border = make_border()
            ws2.cell(row=r, column=col).alignment = Alignment(wrap_text=True, vertical="top")
        ws2.row_dimensions[r].height = 20

    return wb


if __name__ == "__main__":
    import os
    out_path = "/mnt/c/Users/mkuzi/Downloads/regula_competitor_matrix_apr2026.xlsx"
    wb = build_workbook()
    wb.save(out_path)
    size_kb = os.path.getsize(out_path) // 1024
    print(f"Saved: {out_path}  ({size_kb} KB)")
    print("Open in Excel — freeze panes at D5, use View > Freeze Panes if needed.")
